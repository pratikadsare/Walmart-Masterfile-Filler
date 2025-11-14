# streamlit_app.py
# Walmart Masterfile Filler — Horizontal Mapping with High‑Fidelity Header Preview (Rows 2–6, no inner scroll)
# - Styled preview shows rows 2–6 "as-is" (merged cells, fills, fonts/alignment approximated) without its own scrollbar
# - Horizontal mapping row (one dropdown per template column) directly beneath the preview in the same section
# - Supports composite (multi-row) headers for mapping
# - Preserves macros/formatting/validations on export; writes from row 7 (configurable)
# - Auto-save mapping on sheet switch, fuzzy/exact automap, one workbook importer (PCSE/TIC), duplicate highlighting

import re
from io import BytesIO
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# =============================
# Page config & hero
# =============================
st.set_page_config(page_title="Walmart Masterfile Filler", layout="wide")
st.markdown(
    """
    <div style="text-align:center; margin-top:-0.5rem; margin-bottom:0.5rem;">
      <h1 style="margin-bottom:0.25rem;">Walmart Masterfile Filler</h1>
      <p style="margin-top:0; font-size:1.05rem;">Empowering Innovation ⏐ Maximizing Performance</p>
    </div>
    """,
    unsafe_allow_html=True,
)
st.divider()

# =============================
# Canonical target sheets (normalized key -> display name)
# =============================
TARGET_SHEETS_CANON = {
    "productcontentandsiteexp": "Product Content And Site Exp",
    "tradeitemconfigurations": "Trade Item Configurations",
}
CANON_ORDER = ["productcontentandsiteexp", "tradeitemconfigurations"]

# Synonyms for mapping workbook tab detection (import)
MAPPING_TAB_SYNONYMS = {
    "productcontentandsiteexp": ["product content and site exp", "product site exp", "pcse"],
    "tradeitemconfigurations": ["trade item configurations", "trade item", "tic"],
}

# Defaults
DEFAULT_HEADER_ROWS = [5]   # you can change in UI (e.g., 3,4,5 for composite headers)
DEFAULT_DATA_START_ROW = 7  # write from row 7 by default

# Highlight style for duplicate cells (in exported Excel)
YELLOW_DUP_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# =============================
# Utilities
# =============================
def _norm_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower().replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", s)

def _parse_rows_input(s: str, default: List[int]) -> List[int]:
    try:
        vals = [int(x.strip()) for x in s.split(",") if x.strip()]
        vals = sorted(set(v for v in vals if v > 0))
        return vals if vals else default
    except Exception:
        return default

def _enforce_extension(filename: str, is_xlsm: bool) -> str:
    if not filename:
        return "filled_template.xlsm" if is_xlsm else "filled_template.xlsx"
    filename = filename.strip()
    if is_xlsm and not filename.lower().endswith(".xlsm"):
        filename = re.sub(r"\.(xlsx|xls)$", "", filename, flags=re.IGNORECASE) + ".xlsm"
    if not is_xlsm and not filename.lower().endswith(".xlsx"):
        filename = re.sub(r"\.(xlsm|xls)$", "", filename, flags=re.IGNORECASE) + ".xlsx"
    return filename

def _find_target_sheets(actual_names: List[str]) -> Dict[str, str]:
    present: Dict[str, str] = {}
    norm_actual = { _norm_key(n): n for n in actual_names }
    for canon_norm, _display in TARGET_SHEETS_CANON.items():
        if canon_norm in norm_actual:
            present[canon_norm] = norm_actual[canon_norm]
            continue
        for n_norm, real in norm_actual.items():
            if canon_norm in n_norm or n_norm in canon_norm:
                present[canon_norm] = real
                break
    return present

# ---------- Excel reading helpers ----------
def _get_cell_value_with_merge(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value not in (None, ""):
        return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return None

def _extract_headers_composite(ws, header_rows: List[int]) -> Tuple[List[str], Dict[int, int], Dict[str, List[int]]]:
    """
    Build composite headers by combining texts in multiple header rows for each column.
      - headers: ordered list of composite header strings (for columns with any header text)
      - pos_to_col: mapping 1-based header position -> column index
      - norm_to_cols: mapping normalized composite header -> list of actual column indexes
    """
    headers: List[str] = []
    pos_to_col: Dict[int, int] = {}
    norm_to_cols: Dict[str, List[int]] = {}
    max_col = ws.max_column or 0

    for col_idx in range(1, max_col + 1):
        parts: List[str] = []
        for r in header_rows:
            v = _get_cell_value_with_merge(ws, r, col_idx)
            if v is not None and str(v).strip() != "":
                parts.append(str(v).strip())
        if not parts:
            continue
        composite = " | ".join(parts)
        headers.append(composite)
        pos = len(headers)
        pos_to_col[pos] = col_idx
        norm = _norm_key(composite)
        norm_to_cols.setdefault(norm, []).append(col_idx)

    return headers, pos_to_col, norm_to_cols

# ---------- Color & size approximation for HTML preview ----------
THEME_COLOR_MAP = {
    0: "#FFFFFF", 1: "#000000", 2: "#EEECE1", 3: "#1F497D",
    4: "#4F81BD", 5: "#C0504D", 6: "#9BBB59", 7: "#8064A2", 8: "#4BACC6", 9: "#F79646",
}
INDEXED_COLOR_FALLBACK = {64: "#FFFFFF", 9: "#000000"}

def _apply_tint_to_hex(hex_color: str, tint: float) -> str:
    try:
        h = hex_color.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        def adj(ch):
            if tint < 0:
                return int(ch * (1.0 + tint))
            return int(ch + (255 - ch) * tint)
        r = min(255, max(0, adj(r)))
        g = min(255, max(0, adj(g)))
        b = min(255, max(0, adj(b)))
        return f"#{r:02X}{g:02X}{b:02X}"
    except Exception:
        return hex_color

def _excel_color_to_hex(color) -> Optional[str]:
    if color is None:
        return None
    try:
        typ = getattr(color, "type", None)
        if typ == "rgb" and getattr(color, "rgb", None):
            return "#" + color.rgb[-6:]
        if typ == "theme" and getattr(color, "theme", None) is not None:
            base = THEME_COLOR_MAP.get(color.theme)
            tint = getattr(color, "tint", 0.0) or 0.0
            if base:
                return _apply_tint_to_hex(base, float(tint))
            return base
        if typ == "indexed" and getattr(color, "indexed", None) is not None:
            return INDEXED_COLOR_FALLBACK.get(color.indexed)
    except Exception:
        return None
    return None

def _excel_col_width_to_pixels(width: Optional[float]) -> int:
    if width is None:
        width = 8.43
    try:
        return max(32, int(round((width + 0.72) * 7)))
    except Exception:
        return 64

def _excel_row_height_to_pixels(height: Optional[float]) -> int:
    if height is None:
        height = 15.0
    try:
        return max(16, int(round(height * 96.0 / 72.0)))
    except Exception:
        return 18

def _build_header_preview_html(ws, start_row: int = 2, num_rows: int = 5, max_cols: Optional[int] = None) -> Tuple[str, int]:
    """Return (html, exact_pixel_height) for rows start_row..start_row+num_rows-1."""
    max_col = ws.max_column or 0
    if max_cols is not None:
        max_col = min(max_col, max_cols)

    # Precompute merged ranges
    top_left_to_span = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        rowspan = rng.max_row - rng.min_row + 1
        colspan = rng.max_col - rng.min_col + 1
        top_left_to_span[(rng.min_row, rng.min_col)] = (rowspan, colspan)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if not (r == rng.min_row and c == rng.min_col):
                    covered.add((r, c))

    # Column widths
    col_widths_px = []
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width
        col_widths_px.append(_excel_col_width_to_pixels(w))

    # Total height
    rows = range(start_row, start_row + num_rows)
    total_height_px = sum(_excel_row_height_to_pixels(ws.row_dimensions[r].height) for r in rows) + 12  # padding

    # Build HTML table
    html_parts = []
    html_parts.append(
        """
        <div style="max-width:100%; overflow-x:auto; border:1px solid #ddd; border-radius:6px; background:#fff;">
        <table style="border-collapse:collapse; font-family:Arial, Helvetica, sans-serif; font-size:12px; width:max-content;">
        <colgroup>
        """
    )
    for px in col_widths_px:
        html_parts.append(f'<col style="width:{px}px;">')
    html_parts.append("</colgroup><tbody>")

    for r in rows:
        h_px = _excel_row_height_to_pixels(ws.row_dimensions[r].height)
        html_parts.append(f'<tr style="height:{h_px}px;">')
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)

            # Span if merged top-left
            rs, cs = 1, 1
            if (r, c) in top_left_to_span:
                rs_all, cs = top_left_to_span[(r, c)]
                # clamp rowspan to our preview window
                rs = max(1, min(rs_all, start_row + num_rows - r))

            # Background
            bg = None
            fill = getattr(cell, "fill", None)
            if isinstance(fill, PatternFill) and getattr(fill, "fill_type", None) == "solid":
                color_obj = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
                bg = _excel_color_to_hex(color_obj)

            # Font & alignment
            font = cell.font
            bold = "bold" if getattr(font, "bold", False) else "normal"
            italic = "italic" if getattr(font, "italic", False) else "normal"
            underline = getattr(font, "underline", None)
            txt_dec = "underline" if underline and underline != "none" else "none"
            fcolor = None
            try:
                fcolor = _excel_color_to_hex(font.color) if font.color is not None else None
            except Exception:
                fcolor = None

            al = cell.alignment
            hal = getattr(al, "horizontal", None) or "left"
            val = getattr(al, "vertical", None) or "center"
            wrap = getattr(al, "wrap_text", False)

            styles = [
                "border:1px solid #e0e0e0",
                f"text-align:{hal}",
                f"vertical-align:{'middle' if val=='center' else val}",
                f"font-weight:{bold}",
                f"font-style:{italic}",
                f"text-decoration:{txt_dec}",
                "padding:4px 6px",
                f"white-space:{'normal' if wrap else 'nowrap'}",
            ]
            if fcolor:
                styles.append(f"color:{fcolor}")
            if bg:
                styles.append(f"background-color:{bg}")

            value = _get_cell_value_with_merge(ws, r, c)
            text = "" if value in (None, "") else str(value)
            text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

            span_attrs = ""
            if rs > 1:
                span_attrs += f' rowspan="{rs}"'
            if cs > 1:
                span_attrs += f' colspan="{cs}"'

            html_parts.append(f'<td{span_attrs} style="{";".join(styles)}">{text}</td>')
        html_parts.append("</tr>")
    html_parts.append("</tbody></table></div>")
    return "".join(html_parts), total_height_px

# ---------- Widget key helpers ----------
def _map_key(sheet_key: str, templ_norm: str, pos: int) -> str:
    return f"map_{sheet_key}_{pos}_{templ_norm}"

def _tick_key(sheet_key: str, templ_norm: str, pos: int) -> str:
    return f"tick_{sheet_key}_{pos}_{templ_norm}"

def _iter_template_positions(sheet_key: str, templ_headers: List[str]):
    for idx, t in enumerate(templ_headers, start=1):
        templ_norm = _norm_key(t)
        yield t, templ_norm, idx, _map_key(sheet_key, templ_norm, idx), _tick_key(sheet_key, templ_norm, idx)

def _touch_tick(sheet_key: str, templ_norm: str, pos: int):
    st.session_state["last_edit_tick"] = st.session_state.get("last_edit_tick", 0) + 1
    st.session_state[_tick_key(sheet_key, templ_norm, pos)] = st.session_state["last_edit_tick"]

def _build_live_mapping_for_sheet(sheet_key: str, templ_headers: List[str]) -> List[Dict[str, str]]:
    records = []
    for t, templ_norm, pos, mkey, tkey in _iter_template_positions(sheet_key, templ_headers):
        raw = st.session_state.get(mkey, "") or ""
        if raw:
            records.append({
                "template_header": t,
                "template_norm": templ_norm,
                "templ_pos": pos,
                "raw_header": raw,
                "tick": st.session_state.get(tkey, 0),
            })
    return records

def _commit_current_sheet_mapping(sheet_key: str, templ_headers: List[str]):
    st.session_state.setdefault("saved_mapping", {})
    st.session_state["saved_mapping"][sheet_key] = _build_live_mapping_for_sheet(sheet_key, templ_headers)

def _hydrate_live_from_saved(sheet_key: str, templ_headers: List[str]):
    saved = st.session_state.get("saved_mapping", {})
    recs = saved.get(sheet_key, [])
    by_pos = { r.get("templ_pos"): r.get("raw_header") for r in recs if r.get("templ_pos") is not None }
    for t, templ_norm, pos, mkey, _ in _iter_template_positions(sheet_key, templ_headers):
        if not st.session_state.get(mkey, "") and pos in by_pos:
            st.session_state[mkey] = by_pos[pos]

# ---------- Automap ----------
def _auto_map_exact(sheet_key: str, templ_headers: List[str], raw_headers: List[str]):
    norm_idx = { _norm_key(r): r for r in raw_headers }
    for t, templ_norm, pos, mkey, _ in _iter_template_positions(sheet_key, templ_headers):
        if st.session_state.get(mkey, ""):
            continue
        match = norm_idx.get(templ_norm)
        if not match:
            tokens = [tok.strip() for tok in t.split("|") if tok.strip()]
            if tokens:
                match = norm_idx.get(_norm_key(tokens[-1]))
        if match:
            st.session_state[mkey] = match
            _touch_tick(sheet_key, templ_norm, pos)

def _auto_map_fuzzy(sheet_key: str, templ_headers: List[str], raw_headers: List[str], threshold: int = 80):
    thr = max(0, min(100, threshold)) / 100.0
    raw_norms = [(r, _norm_key(r)) for r in raw_headers]
    for t, templ_norm, pos, mkey, _ in _iter_template_positions(sheet_key, templ_headers):
        if st.session_state.get(mkey, ""):
            continue
        best_raw, best_score = None, 0.0
        for r, rnorm in raw_norms:
            score = SequenceMatcher(None, templ_norm, rnorm).ratio()
            if score > best_score:
                best_score, best_raw = score, r
        if (best_raw is None) or (best_score < thr):
            tokens = [tok.strip() for tok in t.split("|") if tok.strip()]
            if tokens:
                last_norm = _norm_key(tokens[-1])
                for r, rnorm in raw_norms:
                    score = SequenceMatcher(None, last_norm, rnorm).ratio()
                    if score > best_score:
                        best_score, best_raw = score, r
        if best_raw and best_score >= thr:
            st.session_state[mkey] = best_raw
            _touch_tick(sheet_key, templ_norm, pos)

# ---------- Import/Export ----------
def _header_match(import_templ: str, display_header: str) -> bool:
    a = _norm_key(import_templ)
    b = _norm_key(display_header)
    if a == b:
        return True
    tokens = re.split(r"\s*\|\s*|\s*>\s*", display_header)
    for tok in tokens:
        if _norm_key(tok) == a:
            return True
    return False

def _import_mapping_df_for_sheet(df: pd.DataFrame,
                                 target_sheet_key: str,
                                 templ_headers_by_sheet: Dict[str, List[str]],
                                 raw_headers: List[str]) -> int:
    cols = { _norm_key(c): c for c in df.columns }
    tcol = next((cols[c] for c in ["template","templateheader","templ","target","targetheader"] if c in cols), None)
    rcol = next((cols[c] for c in ["raw","rawheader","source","sourceheader"] if c in cols), None)
    if not tcol or not rcol:
        return 0

    templ_headers = templ_headers_by_sheet.get(target_sheet_key, [])
    raw_set = set(raw_headers)
    applied = 0
    for _, row in df.iterrows():
        templ_in = str(row[tcol]).strip() if pd.notna(row[tcol]) else ""
        raw = str(row[rcol]).strip() if pd.notna(row[rcol]) else ""
        if not templ_in or not raw or raw not in raw_set:
            continue
        assigned = False
        for t, templ_norm, pos, mkey, _ in _iter_template_positions(target_sheet_key, templ_headers):
            if _header_match(templ_in, t) and not st.session_state.get(mkey, ""):
                st.session_state[mkey] = raw
                _touch_tick(target_sheet_key, templ_norm, pos)
                applied += 1
                assigned = True
                break
        if assigned:
            continue
        for t, templ_norm, pos, mkey, _ in _iter_template_positions(target_sheet_key, templ_headers):
            if _header_match(templ_in, t):
                st.session_state[mkey] = raw
                _touch_tick(target_sheet_key, templ_norm, pos)
                applied += 1
                break
    return applied

def _export_mapping_df(present_sheets: Dict[str, str],
                       templ_headers_by_sheet: Dict[str, List[str]],
                       pos_to_col_by_sheet: Dict[str, Dict[int, int]],
                       only_sheet_key: Optional[str] = None) -> pd.DataFrame:
    rows = []
    for sk in _ordered_keys(present_sheets):
        if only_sheet_key and sk != only_sheet_key:
            continue
        actual_name = present_sheets[sk]
        templ_headers = templ_headers_by_sheet.get(sk, [])
        pos_to_col = pos_to_col_by_sheet.get(sk, {})
        for t, templ_norm, pos, mkey, _ in _iter_template_positions(sk, templ_headers):
            v = st.session_state.get(mkey, "")
            col_letter = get_column_letter(pos_to_col.get(pos, 0)) if pos_to_col.get(pos, 0) else ""
            rows.append({"Sheet": TARGET_SHEETS_CANON.get(sk, actual_name), "Position": pos, "Column": col_letter, "Template": t, "Raw": v})
    return pd.DataFrame(rows)

# ---------- Duplicate RAW selection handling ----------
def _resolve_duplicate_raw_mappings(records: List[Dict[str, str]], auto_resolve: bool) -> Tuple[List[Dict[str, str]], List[str]]:
    bucket: Dict[str, List[Dict[str, str]]] = {}
    for r in records:
        bucket.setdefault(r["raw_header"], []).append(r)
    dups = [raw for raw, lst in bucket.items() if len(lst) > 1]
    if not dups:
        return records, []
    if auto_resolve:
        resolved = []
        for raw, lst in bucket.items():
            if len(lst) == 1:
                resolved.append(lst[0])
            else:
                keep = sorted(lst, key=lambda x: x.get("tick", 0))[-1]
                resolved.append(keep)
        return resolved, []
    else:
        return records, dups

# ---------- Writing ----------
def _write_sheet_data(ws, mapping: List[Dict[str, str]],
                      header_rows: List[int],
                      start_row: int,
                      raw_df: pd.DataFrame,
                      dup_headers_to_highlight: List[str]) -> Tuple[int, List[str]]:
    headers, pos_to_col, norm_to_cols = _extract_headers_composite(ws, header_rows)

    write_plan: List[Tuple[int, str]] = []
    missing: List[str] = []
    for m in mapping:
        pos = m.get("templ_pos")
        if isinstance(pos, int) and pos in pos_to_col:
            col_idx = pos_to_col[pos]
            write_plan.append((col_idx, m["raw_header"]))
        else:
            norm = _norm_key(m["template_header"])
            if norm in norm_to_cols and len(norm_to_cols[norm]) > 0:
                write_plan.append((norm_to_cols[norm][0], m["raw_header"]))
            else:
                missing.append(m["template_header"])

    nrows = len(raw_df)
    for i in range(nrows):
        excel_row = start_row + i
        for col_idx, raw_name in write_plan:
            val = raw_df.iloc[i][raw_name] if raw_name in raw_df.columns else None
            ws.cell(row=excel_row, column=col_idx, value=val)

    # Highlight duplicates in the chosen columns (matched by composite header normalization)
    dup_norms = [_norm_key(x) for x in dup_headers_to_highlight if str(x).strip()]
    for want_norm in dup_norms:
        match_cols = set()
        for norm_h, cols in norm_to_cols.items():
            if want_norm == norm_h or want_norm in norm_h:
                match_cols.update(cols)
        for cidx in match_cols:
            counts: Dict[str, int] = {}
            for i in range(nrows):
                v = ws.cell(row=start_row + i, column=cidx).value
                if v is None or str(v).strip() == "":
                    continue
                key = str(v)
                counts[key] = counts.get(key, 0) + 1
            dup_values = {k for k, c in counts.items() if c > 1}
            if not dup_values:
                continue
            for i in range(nrows):
                cell = ws.cell(row=start_row + i, column=cidx)
                v = cell.value
                if v is None:
                    continue
                if str(v) in dup_values:
                    cell.fill = YELLOW_DUP_FILL
    return nrows, missing

def _build_output_bytes(template_bytes: bytes,
                        template_is_xlsm: bool,
                        sheets_to_process: Dict[str, str],
                        templ_headers_by_sheet: Dict[str, List[str]],
                        header_rows: List[int],
                        raw_df: pd.DataFrame,
                        auto_resolve_dupe_mappings: bool,
                        dup_columns_to_highlight: List[str],
                        saved_mapping_store: Optional[Dict[str, List[Dict[str, str]]]] = None) -> bytes:
    bio_in = BytesIO(template_bytes)
    wb = load_workbook(bio_in, read_only=False, keep_vba=template_is_xlsm, data_only=False)

    for sk, actual_name in sheets_to_process.items():
        ws = wb[actual_name]
        if saved_mapping_store and sk in saved_mapping_store:
            mapping_records = list(saved_mapping_store[sk])
        else:
            mapping_records = _build_live_mapping_for_sheet(sk, templ_headers_by_sheet.get(sk, []))

        mapping_resolved, dup_raw = _resolve_duplicate_raw_mappings(mapping_records, auto_resolve_dupe_mappings)
        if dup_raw:
            raise ValueError(
                f"Duplicate RAW column selections for sheet '{TARGET_SHEETS_CANON.get(sk, actual_name)}': {', '.join(sorted(set(dup_raw)))}. "
                "Turn ON auto‑resolve or change your selections."
            )

        _write_sheet_data(
            ws=ws,
            mapping=mapping_resolved,
            header_rows=header_rows,
            start_row=st.session_state.get("data_start_row", DEFAULT_DATA_START_ROW),
            raw_df=raw_df,
            dup_headers_to_highlight=dup_columns_to_highlight,
        )

    bio_out = BytesIO()
    wb.save(bio_out)  # preserves formatting, heights, colors, validations; keep_vba preserves macros
    return bio_out.getvalue()

# ---------- Ordering / matching helpers ----------
def _ordered_keys(present_sheets: Dict[str, str]) -> List[str]:
    ordered = [k for k in CANON_ORDER if k in present_sheets]
    ordered += [k for k in present_sheets.keys() if k not in ordered]
    return ordered

def _match_mapping_tab(sheet_names: List[str], target_key: str) -> Optional[str]:
    candidates = [(_norm_key(n), n) for n in sheet_names]
    targets = [target_key] + MAPPING_TAB_SYNONYMS.get(target_key, [])
    targets_norm = [_norm_key(x) for x in targets]
    for norm, name in candidates:
        if norm in targets_norm:
            return name
    for norm, name in candidates:
        for t in targets_norm:
            if t in norm or norm in t:
                return name
    return None

# =============================
# Session bootstrap
# =============================
for k, v in [
    ("template_bytes", None),
    ("template_ext", None),
    ("present_sheets", {}),
    ("templ_headers_by_sheet", {}),
    ("pos_to_col_by_sheet", {}),
    ("templ_header_sigs", {}),
    ("raw_headers", []),
    ("raw_sig", ""),
    ("raw_prev_headers", []),
    ("current_sheet_key", None),
    ("last_edit_tick", 0),
    ("download_payload", None),
    ("saved_mapping", {}),
    ("header_rows", DEFAULT_HEADER_ROWS),
    ("data_start_row", DEFAULT_DATA_START_ROW),
]:
    st.session_state.setdefault(k, v)

# =============================
# Preserve scroll position
# =============================
st.components.v1.html(
    """
    <script>
    const KEY = "scrollY";
    window.addEventListener("beforeunload", () => { sessionStorage.setItem(KEY, window.scrollY); });
    window.addEventListener("load", () => {
      const y = sessionStorage.getItem(KEY);
      if (y !== null) window.scrollTo(0, parseFloat(y));
    });
    </script>
    """,
    height=0,
)

# =============================
# Tabs
# =============================
tab1, tab2, tab3 = st.tabs([
    "Upload Masterfile Template",
    "Upload Raw / Onboarding Data",
    "Mapping (Horizontal) & Download",
])

# -----------------------------
# Tab 1: Upload Template
# -----------------------------
with tab1:
    st.markdown("#### Step 1 of 3 — Upload Masterfile Template (.xlsx / .xlsm)")

    c1, c2 = st.columns((2, 1))
    with c1:
        hdr_str_default = ",".join(str(x) for x in st.session_state.get("header_rows", DEFAULT_HEADER_ROWS))
        hdr_str = st.text_input(
            "Header rows to combine (comma‑separated, top→bottom)",
            value=hdr_str_default,
            help="Example: 3,4,5 — texts in those rows (same column) are joined to form a composite header."
        )
        st.session_state["header_rows"] = _parse_rows_input(hdr_str, DEFAULT_HEADER_ROWS)
    with c2:
        dsr = st.number_input("Data start row", min_value=2, step=1, value=st.session_state.get("data_start_row", DEFAULT_DATA_START_ROW))
        st.session_state["data_start_row"] = int(dsr)

    tpl = st.file_uploader("Upload template", type=["xlsx", "xlsm"], key="template_uploader")

    do_rescan = False
    if st.session_state.get("template_bytes"):
        do_rescan = st.button("Re‑scan headers using current rows")

    if tpl is not None or do_rescan:
        if tpl is not None:
            raw_bytes = tpl.read()
            st.session_state["template_bytes"] = raw_bytes
            st.session_state["template_ext"] = "xlsm" if tpl.name.lower().endswith(".xlsm") else "xlsx"
        else:
            raw_bytes = st.session_state["template_bytes"]

        is_xlsm = st.session_state["template_ext"] == "xlsm"
        wb = load_workbook(BytesIO(raw_bytes), read_only=False, keep_vba=is_xlsm, data_only=False)
        all_names = wb.sheetnames
        present = _find_target_sheets(all_names)

        templ_headers_by_sheet: Dict[str, List[str]] = {}
        pos_to_col_by_sheet: Dict[str, Dict[int, int]] = {}
        templ_header_sigs: Dict[str, str] = {}

        for sk, actual in present.items():
            ws = wb[actual]
            headers, pos_to_col, _ = _extract_headers_composite(ws, st.session_state["header_rows"])
            templ_headers_by_sheet[sk] = headers
            pos_to_col_by_sheet[sk] = pos_to_col
            templ_header_sigs[sk] = "|".join(headers)

        prev_sigs = st.session_state.get("templ_header_sigs", {})
        st.session_state["present_sheets"] = present
        st.session_state["templ_headers_by_sheet"] = templ_headers_by_sheet
        st.session_state["pos_to_col_by_sheet"] = pos_to_col_by_sheet
        st.session_state["templ_header_sigs"] = templ_header_sigs

        # Clear live & saved mapping only for sheets whose header signature changed
        for sk, sig in templ_header_sigs.items():
            if prev_sigs.get(sk) != sig:
                p_map = f"map_{sk}_"
                p_tick = f"tick_{sk}_"
                for key in list(st.session_state.keys()):
                    if key.startswith(p_map) or key.startswith(p_tick):
                        del st.session_state[key]
                if sk in st.session_state["saved_mapping"]:
                    del st.session_state["saved_mapping"][sk]

        ordered = _ordered_keys(present)
        if ordered and st.session_state.get("current_sheet_key") not in present:
            st.session_state["current_sheet_key"] = ordered[0]

        st.success("Template loaded & headers scanned.")
        if present:
            st.caption(f"Composite headers from rows: {', '.join(str(r) for r in st.session_state['header_rows'])}. Data will be written from row {st.session_state['data_start_row']}.")
            st.write("Detected target sheets:")
            for sk in _ordered_keys(present):
                st.write(f"- **{TARGET_SHEETS_CANON.get(sk, present[sk])}** (actual: `{present[sk]}`)")
        else:
            st.warning("No target sheets found. Expected PCSE and/or TIC.")

# -----------------------------
# Tab 2: Upload Raw
# -----------------------------
with tab2:
    st.markdown("#### Step 2 of 3 — Upload Raw / Onboarding Data (.csv / .xlsx)")
    raw_file = st.file_uploader("Upload raw/onboarding file", type=["csv", "xlsx"], key="raw_uploader")

    raw_df = None
    if raw_file is not None:
        if raw_file.name.lower().endswith(".csv"):
            raw_df = pd.read_csv(raw_file)
        else:
            xl = pd.ExcelFile(raw_file)
            sheetname = st.selectbox("Choose a sheet", xl.sheet_names, index=0)
            raw_df = xl.parse(sheetname)

        raw_df.columns = [str(c) for c in raw_df.columns]
        new_headers = list(raw_df.columns)
        new_sig = "|".join(new_headers)

        removed = set(st.session_state.get("raw_prev_headers", [])) - set(new_headers)
        if removed:
            for sk in st.session_state.get("present_sheets", {}).keys():
                for k in list(st.session_state.keys()):
                    if k.startswith(f"map_{sk}_") and st.session_state.get(k, "") in removed:
                        st.session_state[k] = ""
                if "saved_mapping" in st.session_state and sk in st.session_state["saved_mapping"]:
                    st.session_state["saved_mapping"][sk] = [
                        r for r in st.session_state["saved_mapping"][sk] if r.get("raw_header") not in removed
                    ]

        st.session_state["raw_prev_headers"] = new_headers
        st.session_state["raw_headers"] = new_headers
        st.session_state["raw_sig"] = new_sig
        st.session_state["raw_df_payload"] = raw_df

        st.success(f"Raw data loaded ({len(raw_df)} rows, {len(new_headers)} columns).")
        with st.expander("Preview (first 20 rows)"):
            st.dataframe(raw_df.head(20), use_container_width=True)

# -----------------------------
# Tab 3: Mapping (Horizontal) & Download
# -----------------------------
with tab3:
    st.markdown("#### Step 3 of 3 — Mapping (Horizontal) & Download")

    if not st.session_state.get("template_bytes"):
        st.info("Please upload a masterfile template in tab 1.")
        st.stop()
    if "raw_df_payload" not in st.session_state:
        st.info("Please upload raw/onboarding data in tab 2.")
        st.stop()

    present = st.session_state["present_sheets"]
    templ_headers_by_sheet = st.session_state["templ_headers_by_sheet"]
    pos_to_col_by_sheet = st.session_state.get("pos_to_col_by_sheet", {})
    raw_headers: List[str] = st.session_state["raw_headers"]
    raw_df: pd.DataFrame = st.session_state["raw_df_payload"]

    if not present:
        st.error("Template has no target sheets to map (expected PCSE and/or TIC).")
        st.stop()

    ordered_keys = _ordered_keys(present)
    display_names = [TARGET_SHEETS_CANON.get(sk, present[sk]) for sk in ordered_keys]
    key_by_display = { TARGET_SHEETS_CANON.get(sk, present[sk]): sk for sk in ordered_keys }

    current_key_before = st.session_state.get("current_sheet_key", ordered_keys[0])
    default_index = 0
    if current_key_before in present:
        dname = TARGET_SHEETS_CANON.get(current_key_before, present.get(current_key_before, ""))
        if dname in display_names:
            default_index = display_names.index(dname)

    selected_display = st.radio("Choose target sheet", options=display_names, index=default_index, horizontal=True)
    sheet_key = key_by_display[selected_display]

    # Auto-save previous sheet mapping when switching
    if current_key_before != sheet_key and current_key_before in present:
        _commit_current_sheet_mapping(current_key_before, templ_headers_by_sheet.get(current_key_before, []))
    st.session_state["current_sheet_key"] = sheet_key

    # Prepare mapping headers for this sheet
    is_xlsm = (st.session_state.get("template_ext", "xlsx") == "xlsm")
    wb = load_workbook(BytesIO(st.session_state["template_bytes"]), read_only=False, keep_vba=is_xlsm, data_only=False)
    ws = wb[present[sheet_key]]

    templ_headers = templ_headers_by_sheet.get(sheet_key, [])
    pos_to_col = pos_to_col_by_sheet.get(sheet_key, {})

    # ---------- High-fidelity header preview (rows 2–6) with NO inner scroll ----------
    st.markdown("**Template preview (rows 2–6)**")
    html_preview, exact_height = _build_header_preview_html(ws, start_row=2, num_rows=5, max_cols=None)
    # Render at the computed height; no inner scrolling -> single page scroll
    st.components.v1.html(html_preview, height=exact_height + 16, scrolling=False)

    # Hydrate mapping UI from saved snapshot (only fill blanks)
    _hydrate_live_from_saved(sheet_key, templ_headers)

    # ---------- Horizontal mapping row using Data Editor ----------
    st.markdown("**Mapping row (choose a raw column for each template column)**")
    if not raw_headers:
        st.warning("Upload raw data first to enable mapping.")
    else:
        # Build one-row DataFrame: columns are composite headers (display), values are current selections
        options = [""] + raw_headers
        initial = {}
        col_config = {}
        for t, templ_norm, pos, mkey, _ in _iter_template_positions(sheet_key, templ_headers):
            col_letter = get_column_letter(pos_to_col.get(pos, 0)) if pos_to_col.get(pos, 0) else ""
            label = f"{col_letter} {t}".strip()
            initial[label] = st.session_state.get(mkey, "")
            col_config[label] = st.column_config.SelectboxColumn(
                label=label,
                options=options,
                required=False,
            )
        seed_df = pd.DataFrame([initial]) if initial else pd.DataFrame([{}])

        edited_df = st.data_editor(
            seed_df,
            column_config=col_config,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"map_editor_{sheet_key}",
        )

        # Push edits back to session_state mapping keys
        if isinstance(edited_df, pd.DataFrame) and len(edited_df) >= 1 and len(templ_headers) > 0:
            row0 = edited_df.iloc[0].to_dict()
            # Build reverse lookup from editor label -> template header & pos
            label_to_pos = {}
            for t, templ_norm, pos, mkey, _ in _iter_template_positions(sheet_key, templ_headers):
                col_letter = get_column_letter(pos_to_col.get(pos, 0)) if pos_to_col.get(pos, 0) else ""
                label = f"{col_letter} {t}".strip()
                label_to_pos[label] = (t, templ_norm, pos, mkey)
            for label, val in row0.items():
                if label in label_to_pos:
                    t, templ_norm, pos, mkey = label_to_pos[label]
                    new_val = "" if (val is None or str(val).strip() == "") else str(val)
                    if new_val != st.session_state.get(mkey, ""):
                        st.session_state[mkey] = new_val
                        if new_val:
                            _touch_tick(sheet_key, templ_norm, pos)

    st.markdown("---")
    # Tools row
    c1, c2, c3 = st.columns((1, 1, 1))
    with c1:
        if st.button("Auto‑map (exact)", use_container_width=True):
            _auto_map_exact(sheet_key, templ_headers, raw_headers)
            st.toast("Exact auto‑map applied (blanks only).")
    with c2:
        fuzz_thr = st.slider("Fuzzy threshold", 0, 100, 80, 1)
        if st.button("Auto‑map (fuzzy)", use_container_width=True):
            _auto_map_fuzzy(sheet_key, templ_headers, raw_headers, threshold=fuzz_thr)
            st.toast(f"Fuzzy auto‑map applied at threshold {fuzz_thr} (blanks only).")
    with c3:
        st.caption(" ")

    st.markdown("---")
    st.markdown("**Import mapping workbook (.xlsx with two tabs: PCSE & TIC)**")
    st.caption("Provide one Excel file containing two sheets — one for **Product Content And Site Exp** and one for **Trade Item Configurations**. Each tab must have columns: **Template**, **Raw**.")
    imp_wb = st.file_uploader("Upload mapping workbook (.xlsx)", type=["xlsx"], key="map_workbook")
    if imp_wb is not None:
        try:
            xl = pd.ExcelFile(imp_wb)
            applied_summary = []
            for sk in ordered_keys:
                tab_name = _match_mapping_tab(xl.sheet_names, sk)
                if not tab_name:
                    applied_summary.append(f"{TARGET_SHEETS_CANON.get(sk, present[sk])}: no matching tab found")
                    continue
                df_imp = xl.parse(tab_name)
                cnt = _import_mapping_df_for_sheet(df_imp, sk, templ_headers_by_sheet, raw_headers)
                applied_summary.append(f"{TARGET_SHEETS_CANON.get(sk, present[sk])}: {cnt} rows applied from '{tab_name}'")
            st.success("Import complete:\n- " + "\n- ".join(applied_summary))
        except Exception as e:
            st.error(f"Failed to import mapping workbook: {e}")

    st.markdown("---")
    export_df_all = _export_mapping_df(present, templ_headers_by_sheet, pos_to_col_by_sheet)
    st.download_button(
        "Export mapping (all sheets)",
        data=export_df_all.to_csv(index=False).encode("utf-8"),
        file_name="mapping_export_all.csv",
        mime="text/csv",
        use_container_width=True,
    )
    export_df_sel = _export_mapping_df(present, templ_headers_by_sheet, pos_to_col_by_sheet, only_sheet_key=sheet_key)
    st.download_button(
        "Export mapping (selected sheet)",
        data=export_df_sel.to_csv(index=False).encode("utf-8"),
        file_name=f"mapping_export_{TARGET_SHEETS_CANON.get(sheet_key, sheet_key)}.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.markdown("---")
    auto_resolve = st.checkbox(
        "Auto‑resolve duplicate *Raw* selections on download (latest edit wins)",
        value=True,
        key="auto_resolve_dupe_mappings"
    )
    dup_cols_text = st.text_input(
        "Duplicate columns to highlight (comma‑separated)",
        value="SKU, productId, manufacturerPartNumber"
    )
    dup_cols = [c.strip() for c in dup_cols_text.split(",") if c.strip()]

    suggested_name = "filled_template.xlsm" if is_xlsm else "filled_template.xlsx"
    file_name_input = st.text_input("Output filename", value=suggested_name)
    final_name = _enforce_extension(file_name_input, is_xlsm=is_xlsm)

    b1, b2, b3 = st.columns((1, 1, 1))
    with b1:
        if st.button("Save mapping (selected sheet)", use_container_width=True):
            _commit_current_sheet_mapping(sheet_key, templ_headers)
            st.success(f"Saved mapping for '{TARGET_SHEETS_CANON.get(sheet_key, present.get(sheet_key, sheet_key))}'.")
    with b2:
        if st.button("Save & Build (process both sheets)", use_container_width=True):
            try:
                _commit_current_sheet_mapping(sheet_key, templ_headers)
                payload = _build_output_bytes(
                    template_bytes=st.session_state["template_bytes"],
                    template_is_xlsm=is_xlsm,
                    sheets_to_process=present,
                    templ_headers_by_sheet=templ_headers_by_sheet,
                    header_rows=st.session_state.get("header_rows", DEFAULT_HEADER_ROWS),
                    raw_df=raw_df,
                    auto_resolve_dupe_mappings=auto_resolve,
                    dup_columns_to_highlight=dup_cols,
                    saved_mapping_store=st.session_state.get("saved_mapping", {}),
                )
                st.session_state["download_payload"] = payload
                st.success("File prepared. Use the download button below.")
            except ValueError as ve:
                st.error(str(ve))
            except Exception as e:
                st.error(f"Failed to build file: {e}")
    with b3:
        if st.button("Build (selected sheet only)", use_container_width=True):
            try:
                payload = _build_output_bytes(
                    template_bytes=st.session_state["template_bytes"],
                    template_is_xlsm=is_xlsm,
                    sheets_to_process={sheet_key: present[sheet_key]},
                    templ_headers_by_sheet=templ_headers_by_sheet,
                    header_rows=st.session_state.get("header_rows", DEFAULT_HEADER_ROWS),
                    raw_df=raw_df,
                    auto_resolve_dupe_mappings=auto_resolve,
                    dup_columns_to_highlight=dup_cols,
                    saved_mapping_store=None,
                )
                st.session_state["download_payload"] = payload
                st.success("File prepared. Use the download button below.")
            except ValueError as ve:
                st.error(str(ve))
            except Exception as e:
                st.error(f"Failed to build file: {e}")

    if st.session_state.get("download_payload"):
        st.download_button(
            "Download Excel",
            data=st.session_state["download_payload"],
            file_name=final_name,
            mime="application/vnd.ms-excel" if is_xlsm else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
